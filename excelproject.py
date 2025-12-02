import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ==================== 配置区域 ====================
MANUAL_FILE_PATH = r"E:\Desktop\连仪段_施工数字射线检测数据移交模板.xlsx"  # 人工评判标准文件路径
INTELLIGENT_FILE_PATH = r"E:\Desktop\20251120_143446.xlsx"  # 智能评判结果文件路径
OUTPUT_FILE_PATH = r"E:\Desktop\excelproject\output_new.xlsx"  # 输出文件路径
# ==================================================

class ExcelComparer:
    def __init__(self):
        self.manual_data = None
        self.intelligent_data = None
        self.match_results = []
        self.defect_keywords = ['圆', '条', '未熔合', '未焊透', '裂纹', '内凹', '咬边', '烧穿', '未见']

    def find_header_row(self, df):
        """查找表头所在的行"""
        for idx in range(min(10, len(df))):
            row_values = df.iloc[idx].astype(str).values
            for val in row_values:
                if '焊口编号' in val:
                    return idx
        return 0

    def normalize_column_names(self, df):
        """标准化列名，去除*号和空格"""
        df.columns = [str(col).strip().rstrip('*') for col in df.columns]
        return df

    def load_excel_data(self, file_path, sheet_name='施工检测缺欠信息表'):
        """加载Excel数据"""
        print(f"正在加载文件: {file_path}")

        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, header=None)
        else:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            except:
                df = pd.read_excel(file_path, header=None)

        # 查找表头行
        header_row = self.find_header_row(df)

        # 重新读取，指定表头行
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, header=header_row)
        else:
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
            except:
                df = pd.read_excel(file_path, header=header_row)

        # 标准化列名
        df = self.normalize_column_names(df)

        print(f"成功加载 {len(df)} 条记录")
        return df

    def extract_start_position(self, pos_str):
        """提取起始位置数字，支持范围格式如 10-45, 10~45, 10至45"""
        pos_str = str(pos_str).strip()
        # 匹配范围格式
        match = re.match(r'^(\d+\.?\d*)[\s]*[-~～至到]', pos_str)
        if match:
            return float(match.group(1))
        # 直接转换为数字
        try:
            return float(pos_str)
        except:
            return None

    def extract_end_position(self, pos_str):
        """提取结束位置数字,支持范围格式如 10-45, 10~45, 10至45"""
        pos_str = str(pos_str).strip()
        # 匹配范围格式,提取结束位置
        match = re.match(r'^(\d+\.?\d*)[\s]*[-~～至到][\s]*(\d+\.?\d*)', pos_str)
        if match:
            return float(match.group(2))  # 返回第二个数字(结束位置)
        # 如果没有范围格式,返回None
        return None

    def fuzzy_match_defect_type(self, type1, type2):
        """模糊匹配缺欠类型，基于关键字"""
        type1_str = str(type1)
        type2_str = str(type2)

        type1_keywords = [kw for kw in self.defect_keywords if kw in type1_str]
        type2_keywords = [kw for kw in self.defect_keywords if kw in type2_str]

        # 有共同关键字则匹配成功
        return len(set(type1_keywords) & set(type2_keywords)) > 0

    def compare_data(self):
        """对比两个数据表（优化版）"""
        print("\n开始对比数据...")

        self.match_results = []
        manual_matched = set()
        intelligent_matched = set()

        # 预处理：按焊口编号和缺欠类型关键字建立索引
        print("正在建立索引...")
        manual_index = {}
        for m_idx, m_row in self.manual_data.iterrows():
            weld_id = str(m_row['焊口编号'])
            # 提取缺欠类型的关键字
            defect_keywords = [kw for kw in self.defect_keywords if kw in str(m_row['缺欠性质'])]

            for kw in defect_keywords:
                key = (weld_id, kw)
                if key not in manual_index:
                    manual_index[key] = []
                manual_index[key].append((m_idx, m_row))

        print(f"索引建立完成，开始匹配...")
        processed = 0

        # 遍历智能评判结果的每一条记录
        for i_idx, i_row in self.intelligent_data.iterrows():
            processed += 1
            if processed % 500 == 0:
                print(f"  已处理 {processed}/{len(self.intelligent_data)} 条智能记录...")

            matched = False
            weld_id = str(i_row['焊口编号'])

            # ============================================================
            #  特殊逻辑：人工记录包含 “未见”
            #    同焊口编号 → 自动匹配成功
            # ============================================================
            manual_same_weld = self.manual_data[self.manual_data['焊口编号'].astype(str) == weld_id]
            manual_unseen = manual_same_weld[
                manual_same_weld['缺欠性质'].astype(str).str.contains("未见")
            ]

            if len(manual_unseen) > 0:
                for m_idx in manual_unseen.index:
                    # 每个人工记录只匹配一次
                    if m_idx not in manual_matched:
                        manual_matched.add(m_idx)
                        intelligent_matched.add(i_idx)

                        self.match_results.append({
                            'manual_idx': m_idx,
                            'intelligent_idx': i_idx,
                            'matched': True
                        })

                # 如果该智能记录已经匹配过至少一个"未见"，跳过普通匹配
                if i_idx in intelligent_matched:
                    continue

            # 提取智能记录的缺欠类型关键字
            i_defect_keywords = [kw for kw in self.defect_keywords if kw in str(i_row['缺欠性质'])]

            # 只在匹配的焊口编号和缺欠类型中查找
            for kw in i_defect_keywords:
                key = (weld_id, kw)
                if key not in manual_index:
                    continue

                for m_idx, m_row in manual_index[key]:
                    # 检查起始位置误差（±20mm）
                    try:
                        i_pos = self.extract_start_position(i_row['缺欠起始位置（mm）'])
                        m_pos = self.extract_start_position(m_row['缺欠起始位置（mm）'])

                        if i_pos is not None and m_pos is not None and abs(i_pos - m_pos) <= 20:
                            matched = True
                            manual_matched.add(m_idx)
                            intelligent_matched.add(i_idx)
                            self.match_results.append({
                                'manual_idx': m_idx,
                                'intelligent_idx': i_idx,
                                'matched': True
                            })
                            break
                    except:
                        continue

                if matched:
                    break

            if not matched:
                self.match_results.append({
                    'manual_idx': None,
                    'intelligent_idx': i_idx,
                    'matched': False
                })

        # 记录未匹配的人工评判标准
        for m_idx in range(len(self.manual_data)):
            if m_idx not in manual_matched:
                self.match_results.append({
                    'manual_idx': m_idx,
                    'intelligent_idx': None,
                    'matched': False
                })

        print(f"对比完成！")
        print(f"  人工评判记录: {len(self.manual_data)} 条")
        print(f"  智能评判记录: {len(self.intelligent_data)} 条")
        print(f"  成功匹配: {len(intelligent_matched)} 对")

    def generate_output_file(self, output_path):
        """生成带颜色标记的输出文件"""
        print(f"\n正在生成输出文件: {output_path}")

        new_data = []
        added_manual_indices = set()

        # 先获取智能表中出现的焊口编号集合
        intelligent_weld_set = set(self.intelligent_data['焊口编号'].astype(str).unique())

        # 获取人工表中出现的焊口编号集合
        manual_weld_set = set(self.manual_data['焊口编号'].astype(str).unique())

        # ============================================================
        # 特殊情况②：智能表有但人工表没有的焊口，需要添加占位记录
        # ============================================================
        welds_only_in_intelligent = intelligent_weld_set - manual_weld_set

        # 为这些焊口创建占位记录（不参与匹配，只用于显示）
        placeholder_records = {}
        for weld_id in welds_only_in_intelligent:
            placeholder_records[weld_id] = {
                '评判类型': '人工评判',
                '焊口编号': weld_id,
                '缺陷性质': '原评未评',
                '起始位置': '',
                '结束位置': '',
                '点数/长度': '',
                '截图': '',
                '级别': '',
                '_color': 'white',  # 不做颜色处理
                '_is_placeholder': True  # 标记为占位记录
            }

        # 添加所有记录
        for result in self.match_results:
            # 只在第一次遇到该人工记录时才添加,且焊口编号在智能表中出现过
            if result['manual_idx'] is not None and result['manual_idx'] not in added_manual_indices:
                m_row = self.manual_data.iloc[result['manual_idx']]

                # 检查焊口编号是否在智能表中出现
                if str(m_row['焊口编号']) not in intelligent_weld_set:
                    continue  # 跳过不在智能表中的人工记录

                # 处理结束位置
                ending_pos_from_table = m_row.get('缺欠结束位置（mm）', '')
                start_pos_str = str(m_row['缺欠起始位置（mm）'])

                # 如果表中有明确的结束位置值,直接使用
                if ending_pos_from_table and str(ending_pos_from_table).strip() and str(
                        ending_pos_from_table).strip() != 'nan':
                    final_ending_pos = str(ending_pos_from_table)
                else:
                    # 表中没有结束位置,尝试从起始位置字符串中提取
                    extracted_end = self.extract_end_position(start_pos_str)
                    if extracted_end is not None:
                        # 起始位置是范围格式(如10-45),使用提取的结束位置
                        final_ending_pos = str(extracted_end)
                    else:
                        # 起始位置不是范围格式,使用起始位置+50作为结束位置
                        start_pos_value = self.extract_start_position(start_pos_str)
                        if start_pos_value is not None:
                            final_ending_pos = str(start_pos_value + 50)
                        else:
                            final_ending_pos = ''

                new_data.append({
                    '评判类型': '人工评判',
                    '焊口编号': str(m_row['焊口编号']),
                    '缺陷性质': str(m_row['缺欠性质']),
                    '起始位置': str(m_row['缺欠起始位置（mm）']),
                    '结束位置': final_ending_pos,
                    '点数/长度': str(m_row['缺欠长度（mm/点）']),
                    '截图': '',
                    '级别': str(m_row['评定等级']),
                    '_color': 'light_green' if result['matched'] else 'red',
                    '_is_placeholder': False
                })
                added_manual_indices.add(result['manual_idx'])

            # 添加智能记录
            if result['intelligent_idx'] is not None:
                i_row = self.intelligent_data.iloc[result['intelligent_idx']]
                weld_id = str(i_row['焊口编号'])

                # ============================================================
                # 如果该焊口只在智能表中存在,先添加占位记录
                # ============================================================
                if weld_id in placeholder_records:
                    new_data.append(placeholder_records[weld_id])
                    del placeholder_records[weld_id]  # 删除已使用的占位记录,避免重复

                color = 'green' if result['matched'] else 'yellow'

                # 未匹配且不是Ⅲ或Ⅳ级的不标黄
                if not result['matched']:
                    level = str(i_row.get('评定等级', ''))
                    if level not in ['Ⅲ', 'Ⅳ', 'III', 'IV']:
                        color = 'white'

                ending_pos = i_row.get('缺欠结束位置(mm)', '')

                new_data.append({
                    '评判类型': '智能评判',
                    '焊口编号': weld_id,
                    '缺陷性质': str(i_row['缺欠性质']),
                    '起始位置': str(i_row['缺欠起始位置（mm）']),
                    '结束位置': str(ending_pos) if ending_pos else '',
                    '点数/长度': str(i_row['缺欠长度（mm/点）']),
                    '截图': '',
                    '级别': str(i_row.get('评定等级', '')),
                    '_color': color,
                    '_is_placeholder': False
                })

        # 创建DataFrame
        df_new = pd.DataFrame(new_data)

        # 先排序（在提取颜色信息之前）
        df_new['焊口编号'] = df_new['焊口编号'].astype(str)
        df_new['_sort_key'] = df_new['评判类型'].map({'人工评判': 0, '智能评判': 1})
        df_new = df_new.sort_values(by=['焊口编号', '_sort_key']).reset_index(drop=True)

        # 排序后再提取颜色信息
        color_info = df_new['_color'].tolist()

        # 删除辅助列
        df_output = df_new.drop(columns=['_color', '_sort_key']).copy()

        # 保存到Excel
        df_output.to_excel(output_path, index=False)

        # 使用openpyxl添加颜色
        wb = load_workbook(output_path)
        ws = wb.active

        fills = {
            'light_green': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),
            'green': PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid'),
            'red': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
            'yellow': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        }

        for idx, color in enumerate(color_info, start=2):
            if color in fills:
                for col in range(1, 9):
                    ws.cell(row=idx, column=col).fill = fills[color]

        wb.save(output_path)
        self.add_statistics_to_excel(output_path)

        print(f"输出文件生成成功！")
        print(f"  总记录数: {len(df_output)} 条")
        print(f"  (人工 {len(self.manual_data)} + 智能 {len(self.intelligent_data)})")

    def contains_defect_keyword(self, defect_type):
        """检查缺欠类型是否包含关键字"""
        defect_str = str(defect_type)
        return any(kw in defect_str for kw in self.defect_keywords)

    def generate_statistics_report(self):
        """生成统计报告并打印到控制台"""
        print("\n" + "=" * 100)
        print("                        详细记录统计")
        print("=" * 100 + "\n")

        # 【一、焊口数量统计】
        print("【一、焊口数量统计】")
        intelligent_welds = set(self.intelligent_data['焊口编号'].astype(str).unique())
        manual_welds = set(self.manual_data['焊口编号'].astype(str).unique())

        common_welds = intelligent_welds & manual_welds
        intelligent_only = intelligent_welds - manual_welds
        manual_only = manual_welds - intelligent_welds

        print(f"  • 智能表焊口总数: {len(intelligent_welds)}")
        print(f"  • 人工表焊口总数: {len(manual_welds)}")
        print(f"  • 智能表独有的焊口数量: {len(intelligent_only)}")
        print(f"  • 人工表独有的焊口数量: {len(manual_only)}")
        print(f"  • 二者同时拥有的焊口数量: {len(common_welds)}\n")

        # 【二、匹配情况统计】
        print("【二、匹配情况统计】")
        matched_intelligent_welds = set()
        for result in self.match_results:
            if result['matched'] and result['intelligent_idx'] is not None:
                i_row = self.intelligent_data.iloc[result['intelligent_idx']]
                matched_intelligent_welds.add(i_row['焊口编号'])

        print(f"  • 智能评判结果表中满足人工评判标准表的焊口数量: {len(matched_intelligent_welds)}\n")

        # 【三、标黄记录统计】
        print("【三、标黄记录统计(智能表中未匹配的Ⅲ、Ⅳ级缺陷)】")
        yellow_welds = set()
        yellow_records = 0
        for result in self.match_results:
            if not result['matched'] and result['intelligent_idx'] is not None:
                i_row = self.intelligent_data.iloc[result['intelligent_idx']]
                level = str(i_row.get('评定等级', ''))
                if level in ['Ⅲ', 'Ⅳ', 'III', 'IV']:
                    yellow_welds.add(i_row['焊口编号'])
                    yellow_records += 1

        print(f"  • 智能评判结果表中标黄的焊口数量: {len(yellow_welds)}")
        print(f"  • 智能评判结果表中标黄的记录总数: {yellow_records}\n")

        # 【四、人工表关键字记录统计】
        print("【四、人工表关键字记录统计】")
        manual_keyword_count = 0
        manual_keyword_details = {kw: 0 for kw in self.defect_keywords}

        for _, row in self.manual_data.iterrows():
            if str(row['焊口编号']) in intelligent_welds:
                defect_type = str(row['缺欠性质'])
                for kw in self.defect_keywords:
                    if kw in defect_type:
                        manual_keyword_count += 1
                        manual_keyword_details[kw] += 1
                        break

        detail_str = '、'.join([f"{kw}-{count}" for kw, count in manual_keyword_details.items() if count > 0])
        print(
            f"  • 人工评判标准表中包含关键字(圆、条、未熔合、未焊透、裂纹、内凹、咬边、烧穿、未见)且焊口编号在智能表中出现过的记录总数: {manual_keyword_count}")
        print(f"    其中包括: {detail_str}\n")

        # 【五、人工表关键字匹配成功统计】
        print("【五、人工表关键字匹配成功统计】")
        manual_matched_keyword_count = 0
        manual_matched_keyword_details = {kw: 0 for kw in self.defect_keywords}
        counted_manual_indices = set()

        for result in self.match_results:
            if result['matched'] and result['manual_idx'] is not None:
                if result['manual_idx'] in counted_manual_indices:
                    continue

                m_row = self.manual_data.iloc[result['manual_idx']]
                defect_type = str(m_row['缺欠性质'])
                for kw in self.defect_keywords:
                    if kw in defect_type:
                        manual_matched_keyword_count += 1
                        manual_matched_keyword_details[kw] += 1
                        counted_manual_indices.add(result['manual_idx'])
                        break

        detail_str = '、'.join([f"{kw}-{count}" for kw, count in manual_matched_keyword_details.items() if count > 0])
        print(
            f"  • 人工评判标准表中包含关键字(圆、条、未熔合、未焊透、裂纹、内凹、咬边、烧穿、未见)且成功与智能评判表匹配的记录总数: {manual_matched_keyword_count}")
        print(f"    其中包括: {detail_str}\n")

        print("=" * 100)

    def add_statistics_to_excel(self, output_path):
        """将统计报告添加到Excel的Sheet2"""
        wb = load_workbook(output_path)

        # 创建新的Sheet2
        if 'Sheet2' in wb.sheetnames:
            ws_stats = wb['Sheet2']
            wb.remove(ws_stats)
        ws_stats = wb.create_sheet('统计报告', 1)

        # 生成统计数据
        intelligent_welds = set(self.intelligent_data['焊口编号'].astype(str).unique())
        manual_welds = set(self.manual_data['焊口编号'].astype(str).unique())

        common_welds = intelligent_welds & manual_welds
        intelligent_only = intelligent_welds - manual_welds
        manual_only = manual_welds - intelligent_welds

        matched_intelligent_welds = set()
        for result in self.match_results:
            if result['matched'] and result['intelligent_idx'] is not None:
                i_row = self.intelligent_data.iloc[result['intelligent_idx']]
                matched_intelligent_welds.add(i_row['焊口编号'])

        yellow_welds = set()
        yellow_records = 0
        for result in self.match_results:
            if not result['matched'] and result['intelligent_idx'] is not None:
                i_row = self.intelligent_data.iloc[result['intelligent_idx']]
                level = str(i_row.get('评定等级', ''))
                if level in ['Ⅲ', 'Ⅳ', 'III', 'IV']:
                    yellow_welds.add(i_row['焊口编号'])
                    yellow_records += 1

        # 第④条统计
        manual_keyword_count = 0
        manual_keyword_details = {kw: 0 for kw in self.defect_keywords}

        for _, row in self.manual_data.iterrows():
            if str(row['焊口编号']) in intelligent_welds:
                defect_type = str(row['缺欠性质'])
                for kw in self.defect_keywords:
                    if kw in defect_type:
                        manual_keyword_count += 1
                        manual_keyword_details[kw] += 1
                        break

        # 第⑤条统计
        manual_matched_keyword_count = 0
        manual_matched_keyword_details = {kw: 0 for kw in self.defect_keywords}
        counted_manual_indices = set()

        for result in self.match_results:
            if result['matched'] and result['manual_idx'] is not None:
                if result['manual_idx'] in counted_manual_indices:
                    continue

                m_row = self.manual_data.iloc[result['manual_idx']]
                defect_type = str(m_row['缺欠性质'])
                for kw in self.defect_keywords:
                    if kw in defect_type:
                        manual_matched_keyword_count += 1
                        manual_matched_keyword_details[kw] += 1
                        counted_manual_indices.add(result['manual_idx'])
                        break

        # 写入Sheet2
        row = 1
        ws_stats.cell(row, 1, "详细记录统计")
        ws_stats.cell(row, 1).font = Font(size=16, bold=True)
        row += 2

        ws_stats.cell(row, 1, "【一、焊口数量统计】")
        ws_stats.cell(row, 1).font = Font(bold=True)
        row += 1

        ws_stats.cell(row, 1, f"• 智能表焊口总数: {len(intelligent_welds)}")
        row += 1
        ws_stats.cell(row, 1, f"• 人工表焊口总数: {len(manual_welds)}")
        row += 1
        ws_stats.cell(row, 1, f"• 智能表独有的焊口数量: {len(intelligent_only)}")
        row += 1
        ws_stats.cell(row, 1, f"• 人工表独有的焊口数量: {len(manual_only)}")
        row += 1
        ws_stats.cell(row, 1, f"• 二者同时拥有的焊口数量: {len(common_welds)}")
        row += 2

        ws_stats.cell(row, 1, "【二、匹配情况统计】")
        ws_stats.cell(row, 1).font = Font(bold=True)
        row += 1
        ws_stats.cell(row, 1, f"• 智能评判结果表中满足人工评判标准表的焊口数量: {len(matched_intelligent_welds)}")
        row += 2

        ws_stats.cell(row, 1, "【三、标黄记录统计(智能表中未匹配的Ⅲ、Ⅳ级缺陷)】")
        ws_stats.cell(row, 1).font = Font(bold=True)
        row += 1
        ws_stats.cell(row, 1, f"• 智能评判结果表中标黄的焊口数量: {len(yellow_welds)}")
        row += 1
        ws_stats.cell(row, 1, f"• 智能评判结果表中标黄的记录总数: {yellow_records}")
        row += 2

        ws_stats.cell(row, 1, "【四、人工表关键字记录统计】")
        ws_stats.cell(row, 1).font = Font(bold=True)
        row += 1
        detail_str1 = '、'.join([f"{kw}-{count}" for kw, count in manual_keyword_details.items() if count > 0])
        ws_stats.cell(row, 1,
                      f"• 人工评判标准表中包含关键字且焊口编号在智能表中出现过的记录总数: {manual_keyword_count}")
        row += 1
        ws_stats.cell(row, 1, f"  其中包括: {detail_str1}")
        row += 2

        ws_stats.cell(row, 1, "【五、人工表关键字匹配成功统计】")
        ws_stats.cell(row, 1).font = Font(bold=True)
        row += 1
        detail_str2 = '、'.join([f"{kw}-{count}" for kw, count in manual_matched_keyword_details.items() if count > 0])
        ws_stats.cell(row, 1,
                      f"• 人工评判标准表中包含关键字且成功与智能评判表匹配的记录总数: {manual_matched_keyword_count}")
        row += 1
        ws_stats.cell(row, 1, f"  其中包括: {detail_str2}")

        # 调整列宽
        ws_stats.column_dimensions['A'].width = 120

        wb.save(output_path)

    def run(self, manual_path, intelligent_path, output_path):
        """执行完整的对比流程"""
        try:
            # 1. 加载数据
            self.manual_data = self.load_excel_data(manual_path)
            self.intelligent_data = self.load_excel_data(intelligent_path)

            # 2. 对比数据
            self.compare_data()

            # 3. 生成输出文件
            self.generate_output_file(output_path)

            # 4. 生成统计报告
            self.generate_statistics_report()

            print("\n所有任务完成！")
            return True

        except Exception as e:
            print(f"\n错误: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

def main():
    """主函数"""
    print("=" * 100)
    print("          施工检测缺欠信息对比系统")
    print("=" * 100 + "\n")

    comparer = ExcelComparer()
    success = comparer.run(
        MANUAL_FILE_PATH,
        INTELLIGENT_FILE_PATH,
        OUTPUT_FILE_PATH
    )

    if success:
        print("\n程序执行成功！")
    else:
        print("\n程序执行失败，请检查错误信息。")

if __name__ == '__main__':
    main()