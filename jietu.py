import os
import traceback
from collections import defaultdict
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image
import numpy as np
'''已知信息： self.excel_path为生成的新excel文件路径     self.dcm_path为dcm底片文件夹路径     self.pix_lenth为像素尺寸 '''
# 截图保存目录(可修改)
TMP_SAVE_DIR = r".\tmp_insert_images"
# 扩展像素量
OUTER_EXPAND = 10
# Excel列名
COL_WELD = "焊口编号"
COL_START = "起始位置"
COL_END = "结束位置"
COL_SCREENSHOT = "截图"

def is_row_red(ws, row_idx):
    """判断Excel行是否标红"""
    for col in range(1, 13):
        cell = ws.cell(row=row_idx, column=col)
        fill = getattr(cell, "fill", None)
        if fill is None:
            continue
        start_color = getattr(fill, "start_color", None)
        if start_color is None:
            continue
        rgb = getattr(start_color, "rgb", None)
        if rgb:
            if "FF0000" in rgb.upper():
                return True
    return False

def find_file_with_extension(directory, base_name):
    """
    在目录中查找指定基础名称的文件,支持.dcm和.DICONDE扩展名(不区分大小写)
    返回完整文件路径,如果未找到返回None
    """
    possible_extensions = ['.dcm', '.DCM', '.diconde', '.DICONDE', '.Diconde']

    for ext in possible_extensions:
        file_path = os.path.join(directory, base_name + ext)
        if os.path.exists(file_path):
            return file_path

    return None

def parse_digits_info(digits_info):
    """
    解析digits_info,提取签字信息
    返回字典: {digit_value: [(x, y)], ...}
    """
    sign_dict = defaultdict(list)

    for item in digits_info:
        digit_val = item['digit']
        center = item['center']
        # center可能是list或tuple
        if isinstance(center, (list, tuple)) and len(center) >= 2:
            x, y = float(center[0]), float(center[1])
            sign_dict[digit_val].append((x, y))

    return sign_dict

def find_sign_pair_for_defect(sign_dict, start_mm, end_mm, magnification):
    """
    寻找能包含缺陷的最小签字对
    返回: ((val_left, x_left, y_left), (val_right, x_right, y_right)) 或 None
    """
    if not sign_dict:
        return None

    # 转换为签字数值
    start_sign = start_mm / magnification
    end_sign = end_mm / magnification

    sign_vals = sorted(sign_dict.keys())#排序，用于寻找最小铅字对

    # 找左签字: 最大的 <= start_sign
    left_candidates = [v for v in sign_vals if v <= start_sign]
    if not left_candidates:
        return None
    val_left = max(left_candidates)

    # 找右签字: 最小的 >= end_sign
    right_candidates = [v for v in sign_vals if v >= end_sign]
    if not right_candidates:
        return None
    val_right = min(right_candidates)

    # 选择occurrence
    # 左签字选最靠右(x最大)
    left_occ = max(sign_dict[val_left], key=lambda p: p[0])
    xl, yl = left_occ

    # 右签字选最靠左(x最小)
    right_occ = min(sign_dict[val_right], key=lambda p: p[0])
    xr, yr = right_occ

    return (val_left, xl, yl), (val_right, xr, yr)

def compute_crop_rect(left_sign, right_sign, seam_top, seam_bottom):
    """
    计算裁剪矩形
    返回: (left, top, right, bottom)
    """
    _, x_left, y_left = left_sign
    _, x_right, y_right = right_sign

    # 横向: 左右扩展
    left = min(x_left, x_right) - OUTER_EXPAND
    right = max(x_left, x_right) + OUTER_EXPAND

    # 竖向: 判断签字位置
    sign_y_mean = (y_left + y_right) / 2.0

    if sign_y_mean > seam_bottom:
        # 签字在焊缝下方
        top = seam_top - OUTER_EXPAND
        bottom = int(max(y_left, y_right) + OUTER_EXPAND)
    elif sign_y_mean < seam_top:
        # 签字在焊缝上方
        top = int(min(y_left, y_right) - OUTER_EXPAND)
        bottom = seam_bottom + OUTER_EXPAND
    else:
        # 签字在焊缝区域内
        top = seam_top - OUTER_EXPAND
        bottom = seam_bottom + OUTER_EXPAND

    return int(left), int(top), int(right), int(bottom)

def run(self):
    # 创建临时保存目录
    os.makedirs(TMP_SAVE_DIR, exist_ok=True)
    # 加载Excel文件
    if not os.path.exists(self.excel_path):
        print(f"错误: Excel文件不存在: {self.excel_path}")
        return
    wb = load_workbook(self.excel_path)
    ws = wb.active
    # 解析表头
    header_row = 1
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            headers[str(val).strip()] = col
    # 检查必需列
    if COL_WELD not in headers or COL_START not in headers or COL_END not in headers or COL_SCREENSHOT not in headers:
        print(f"错误: Excel缺少必需列。找到的列: {list(headers.keys())}")
        return
    col_weld_idx = headers[COL_WELD]
    col_start_idx = headers[COL_START]
    col_end_idx = headers[COL_END]
    col_screenshot_idx = headers[COL_SCREENSHOT]
    total_rows = 0
    red_rows_found = 0
    images_inserted = 0
    print("\n开始遍历Excel记录...")

    # 遍历数据行
    for r in range(header_row + 1, ws.max_row + 1):
        total_rows += 1
        try:
            # 检查是否为红色行，如果未标红，则跳过；如果标红，则继续下面逻辑
            if not is_row_red(ws, r):
                continue

            red_rows_found += 1

            # 获取焊口编号、起始位置、结束位置
            weld_id = str(ws.cell(row=r, column=col_weld_idx).value).strip()
            start_val_raw = ws.cell(row=r, column=col_start_idx).value
            end_val_raw = ws.cell(row=r, column=col_end_idx).value

            # 转换为数字
            try:
                start_mm = float(start_val_raw)
                end_mm = float(end_val_raw)
            except:
                print(f"  Row {r}，焊口 {weld_id}: 失败，原因=起始/结束位置非数字")
                continue

            # 定位底片文件
            file_path = find_file_with_extension(self.dcm_path, weld_id)
            if file_path is None:
                print(f"  Row {r}，焊口 {weld_id}: 失败，原因=未找到底片文件")
                continue

            # 初始化焊缝区域
            self.hanfeng_start = 0
            self.hanfeng_end = 0
            # 调用process获取焊缝边界
            self.process(file_path)
            # 调用识别获取签字和倍数信息
            if self.dataThread.cuda_version_float >= 11.3:
                self.dataThread.start()
                self.dataThread.wait()
            else:
                self.All_Info = [0, [], 10]
                print(f"  Row {r}，焊口 {weld_id}: 失败，原因=CUDA版本过低")
                continue
            # 提取信息
            file_name = self.All_Info[0]
            digits_info = self.All_Info[1]
            digit_multiplier = self.All_Info[2]
            hanfeng_start = self.hanfeng_start
            hanfeng_end = self.hanfeng_end
            xiangsu_chicun = self.pix_lenth
            juzhen = self.yuan_juzhen
            # 检查签字信息
            if not digits_info:
                print(f"  Row {r}，焊口 {weld_id}: 失败，原因=签字信息为空")
                continue

            # 解析签字信息
            sign_dict = parse_digits_info(digits_info)

            # 查找签字对
            sign_pair = find_sign_pair_for_defect(sign_dict, start_mm, end_mm, digit_multiplier)
            if sign_pair is None:
                print(f"  Row {r}，焊口 {weld_id}: 失败，原因=未能找到合适的签字对")
                continue

            left_sign, right_sign = sign_pair

            # 计算裁剪矩形
            left, top, right, bottom = compute_crop_rect(left_sign, right_sign, hanfeng_start, hanfeng_end)

            # 修正到图像边界
            img_h, img_w = juzhen.shape[:2]
            left = max(0, left)
            top = max(0, top)
            right = min(img_w, right)
            bottom = min(img_h, bottom)

            # 检查尺寸
            if right - left < 5 or bottom - top < 5:
                print(f"  Row {r}，焊口 {weld_id}: 失败，原因=裁剪尺寸过小")
                continue

            # 从像素矩阵裁剪图像
            cropped_array = juzhen[top:bottom, left:right]

            # 转换为PIL图像
            if len(cropped_array.shape) == 2:
                # 灰度图
                pil_img = Image.fromarray(cropped_array.astype('uint8'), mode='L')
            else:
                # 彩色图
                pil_img = Image.fromarray(cropped_array.astype('uint8'), mode='RGB')

            # 保存临时图片
            tmp_img_path = os.path.join(TMP_SAVE_DIR, f"{weld_id}_{r}.jpg")
            pil_img.save(tmp_img_path, format="JPEG")

            # 插入Excel（优化版）
            try:
                target_col_letter = get_column_letter(col_screenshot_idx)

                # ① 固定截图列宽为 16
                fixed_col_width = 16
                ws.column_dimensions[target_col_letter].width = fixed_col_width

                # ② 计算列宽像素（Excel 字符宽换算）
                col_width_px = int(fixed_col_width * 7)

                # 原图大小
                img_w, img_h = pil_img.size

                # ③ 按列宽缩放，使图像横向恰好占满单元格
                scale_w = col_width_px / img_w
                new_w = col_width_px
                new_h = int(img_h * scale_w)

                # ④ 计算行高 pt（像素 × 0.75），但行高不能低于默认 15pt
                desired_row_height = new_h * 0.75
                MIN_ROW_HEIGHT_PT = 15
                final_row_height = max(desired_row_height, MIN_ROW_HEIGHT_PT)
                ws.row_dimensions[r].height = final_row_height

                # ⑤ 生成缩放后的图片并覆盖保存
                pil_img_resized = pil_img.resize((new_w, new_h))
                pil_img_resized.save(tmp_img_path)

                # ⑥ 插入图片，并允许随单元格移动/改变大小
                xl_img = XLImage(tmp_img_path)

                anchor_cell = f"{target_col_letter}{r}"
                xl_img.anchor = anchor_cell  # 必须先设置 anchor
                ws.add_image(xl_img)  # 然后再 add

                images_inserted += 1
                print(f"  Row {r}, 焊口 {weld_id}: 插入成功，起始={start_mm}, 结束={end_mm}, 倍数={digit_multiplier}，"
                      f"签字对={left_sign[0], right_sign[0]}, "
                      f"签字坐标={(int(left_sign[1]), int(left_sign[2])), (int(right_sign[1]), int(right_sign[2]))}, "
                      f"裁剪={left, top, right, bottom}")

            except Exception as e_img:
                print(f"  Row {r}，焊口 {weld_id}: 失败，原因=插入图片异常: {str(e_img)}")
                continue

        except Exception as e_row:
            traceback.print_exc()
            print(f"  Row {r}: 失败，原因=处理行时异常: {str(e_row)}")
            continue

    # 保存Excel
    try:
        wb.save(self.excel_path)
        print(f"\n===== 插入图片操作完成 =====")
        print(f"总行数扫描: {total_rows}")
        print(f"检测到红色行: {red_rows_found}")
        print(f"成功插入图片: {images_inserted}")
    except Exception as e_save:
        print(f"错误: 保存Excel时发生异常: {e_save}")
            
'''self.All_Info 的内容是：['C:\\Users\\user\\Desktop\\test\\LYY1T01-AC053-004-Z-X01.DICONDE', [{'center': [71.0, 1058.0], 'digit': 255, 'score': 1.0}, {'center': [423.0, 1024.75], 'digit': 260, 'score': 0.9999995231628418}, {'center': [726.5, 1018.5], 'digit': 5, 'score': 0.999998927116394}, {'center': [1101.0, 1020.5], 'digit': 10, 'score': 0.999431848526001}, {'center': [1836.75, 1020.5], 'digit': 20, 'score': 0.9999991655349731}, {'center': [2762.5, 1012.0], 'digit': 20, 'score': 0.9999916553497314}, {'center': [3132.0, 1008.0], 'digit': 25, 'score': 0.9999977946281433}, {'center': [3490.0, 1012.5], 'digit': 30, 'score': 0.9999971389770508}, {'center': [3867.5, 1018.0], 'digit': 35, 'score': 1.0}, {'center': [4799.0, 1009.0], 'digit': 35, 'score': 1.0}, {'center': [5164.0, 1007.5], 'digit': 40, 'score': 0.9998757243156433}, {'center': [5528.5, 1017.5], 'digit': 45, 'score': 0.9737262725830078}, {'center': [5905.5, 1026.5], 'digit': 50, 'score': 0.9999973177909851}, {'center': [6834.5, 1018.0], 'digit': 50, 'score': 0.9999986290931702}, {'center': [7199.5, 1010.5], 'digit': 55, 'score': 1.0}, {'center': [7564.5, 1010.0], 'digit': 60, 'score': 0.9999990463256836}, {'center': [7940.5, 1013.0], 'digit': 65, 'score': 0.9999999403953552}, {'center': [8866.5, 1004.5], 'digit': 65, 'score': 0.9999998807907104}, {'center': [9244.5, 1006.0], 'digit': 70, 'score': 0.999990701675415}, {'center': [9980.75, 1011.0], 'digit': 80, 'score': 0.9999991059303284}, {'center': [10904.0, 1004.0], 'digit': 80, 'score': 0.9999967813491821}, {'center': [11283.5, 1006.5], 'digit': 85, 'score': 1.0}, {'center': [11655.5, 1019.5], 'digit': 90, 'score': 0.999966025352478}, {'center': [12032.5, 1034.5], 'digit': 95, 'score': 0.9999997615814209}, {'center': [12945.0, 1027.5], 'digit': 95, 'score': 0.9999970197677612}, {'center': [13327.0, 1032.5], 'digit': 100, 'score': 0.9999012351036072}, {'center': [13707.5, 1041.5], 'digit': 105, 'score': 0.9997713565826416}, {'center': [14609.0, 1051.0], 'digit': 105, 'score': 0.9957727789878845}, {'center': [15743.5, 1037.0], 'digit': 120, 'score': 0.9999987483024597}, {'center': [16123.5, 1045.0], 'digit': 125, 'score': 0.9999999403953552}, {'center': [17030.0, 1033.5], 'digit': 125, 'score': 0.9999974370002747}, {'center': [17413.0, 1031.0], 'digit': 130, 'score': 0.9999942779541016}, {'center': [17789.0, 1028.5], 'digit': 135, 'score': 0.9999980330467224}, {'center': [18167.5, 1038.0], 'digit': 140, 'score': 0.9999971985816956}, {'center': [19071.0, 1030.0], 'digit': 140, 'score': 0.9999744296073914}, {'center': [19444.0, 1023.5], 'digit': 145, 'score': 0.9999983906745911}, {'center': [19815.0, 1027.5], 'digit': 150, 'score': 0.999978244304657}, {'center': [20196.5, 1023.0], 'digit': 155, 'score': 0.9999864101409912}, {'center': [21106.25, 1015.5], 'digit': 155, 'score': 0.9983065128326416}, {'center': [21483.0, 1006.5], 'digit': 160, 'score': 0.9999990463256836}, {'center': [21853.0, 1004.5], 'digit': 165, 'score': 0.9999958872795105}, {'center': [22237.0, 1019.0], 'digit': 170, 'score': 0.9999998211860657}, {'center': [23145.0, 1011.0], 'digit': 170, 'score': 0.9999797940254211}, {'center': [23518.0, 1012.5], 'digit': 175, 'score': 1.0}, {'center': [23887.5, 1018.5], 'digit': 180, 'score': 1.0}, {'center': [24271.5, 1024.5], 'digit': 185, 'score': 1.0}, {'center': [25187.0, 1016.5], 'digit': 185, 'score': 0.9999985694885254}, {'center': [25557.5, 1004.0], 'digit': 190, 'score': 0.9999942779541016}, {'center': [25928.5, 1006.0], 'digit': 195, 'score': 0.9999846816062927}, {'center': [26297.25, 1020.5], 'digit': 200, 'score': 0.9997150301933289}, {'center': [27209.0, 1012.5], 'digit': 200, 'score': 0.9999999403953552}, {'center': [27580.0, 1012.0], 'digit': 205, 'score': 1.0}, {'center': [27942.5, 1015.5], 'digit': 210, 'score': 0.9999999403953552}, {'center': [28321.0, 1020.5], 'digit': 215, 'score': 0.8917851448059082}, {'center': [29253.0, 1018.5], 'digit': 215, 'score': 0.9999648928642273}, {'center': [29622.5, 1008.5], 'digit': 220, 'score': 1.0}, {'center': [30000.75, 1015.0], 'digit': 225, 'score': 0.859564483165741}, {'center': [30364.0, 1027.5], 'digit': 230, 'score': 1.0}, {'center': [31289.0, 1020.5], 'digit': 230, 'score': 1.0}, {'center': [31657.75, 1033.5], 'digit': 235, 'score': 1.0}, {'center': [32023.0, 1041.0], 'digit': 240, 'score': 0.9999995231628418}, {'center': [32391.75, 1059.5], 'digit': 245, 'score': 0.9927439093589783}, {'center': [33325.0, 1056.5], 'digit': 245, 'score': 0.9306972026824951}, {'center': [33702.0, 1044.5], 'digit': 250, 'score': 0.9999999403953552}, 
{'center': [34062.0, 1038.5], 'digit': 255, 'score': 0.9999999403953552}, {'center': (34383.5, 1016.0), 'digit': 260, 'score': 0.9999382495880127}, {'center': [35335.0, 1022.0], 'digit': 260, 'score': 0.999997615814209}], 10]'''